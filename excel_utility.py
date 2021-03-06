import openpyxl as exel
import os
from utility_file import score2Grade, gradeToNumeric, stRanking, get_csub_status
from selection_criteria import OfferedProgramme
from xlsx_csv_converter import csvWriter
dir_src = os.path.dirname(__file__)

def writeFilename(filename):
    filename = filename.lower().replace(' ', '_')
    working_file = dir_src + '\\' + filename + '.xlsx'
    return working_file

def create_workbook():
    # A copy of workbook and one single worksheet is produce to the calling functions

    wkbook = exel.Workbook()
    wsheet = wkbook.active
    return wkbook, wsheet

def creatSubjecteHeader(subjectList, filename):
    wkbook, current_sheet = create_workbook()
    num_columns = len(subjectList)

    initial_column = 0
    while initial_column < num_columns:
        current_sheet.cell(row=1, column=initial_column+1).value = subjectList[initial_column]
        initial_column += 1
    wkbook.save(filename)

    
def retrievedColhead(filename):
    wkbook = exel.load_workbook(filename)
    wksheet = wkbook.active

    for row in wksheet.iter_rows(min_row=1, max_row=1,values_only=True):
        return (row, wksheet, wkbook)


def getColnum(colobj, wksheet):
    for cellrows in wksheet.iter_rows():
        for cellrow in cellrows:
            if cellrow.value == colobj:
                return cellrow.column
            

def getUserdata(*args):
    cols, wksheet, wkbook = retrievedColhead(args[1])
    num_student = len(args[0])
    data = args[0]
    programme_max_quota = args[2]
    
    counter = 0
    while counter < num_student:
        student_data = get_each_student(data, counter)
        user = student_data.keys()
        value = student_data.values()
        print('writing ......', list(user)[0], 'data to the external file')
        writeOutData2excel(counter+1, user, value, cols, wksheet, data, programme_max_quota)
        counter += 1
    
    wkbook.save(args[1])
    # print(args[1])
    csvWriter(args[1])
    print('Data completely writing to Excel File, you can \
    check it in the root folder where this code is Execute')
    


def get_each_student(studentList, index):
    return studentList[index]
    

def writeOutData2excel(*args):
    num_row = args[0] + 1
    key = list(args[1])[0]
    value = list(args[2])[0]
    programme_max_quota = args[6]
    cols    = args[3]
    wksheet = args[4]
    original_data = args[5]    
    olevel_sub = value[0]
    gender_status = value[2]
    first_choice_status = value[1][1]
    jamb_sub = value[3]
    prefered_programme = value[4]
    olevelAggrScore = value[5][1]
    posutmeAggrscore = value[6][1]
    utmeAggr_Score = value[7][1]
    final_score = value[8][1]
    admission_status = value[9][1]
    st_age = value[10][2]
    catchemment_status = value[10][0]
    studentID, student_ranking = stRanking(args[1], args[2], original_data)
    math_eng_status = None

    for col in cols:
        if col == 'stunid':
            colId = getColnum(col, wksheet)
            wksheet.cell(row=num_row, column=colId).value = key
        elif col in list(olevel_sub):
            for sub, scr in olevel_sub.items():
                if col == sub:
                    colId = getColnum(col, wksheet)
                    wksheet.cell(row=num_row, column=colId).value = \
                    gradeToNumeric(score2Grade(scr))    # Convert to weighted value
                
                if sub == 'general_mathematic':
                    math_status = math_eng_status = get_csub_status(gradeToNumeric(score2Grade(scr)))
                elif sub == 'english_language': 
                    eng_status = math_eng_status = get_csub_status(gradeToNumeric(score2Grade(scr)))
                    
        elif col in list(jamb_sub):
            for sub, scr in jamb_sub.items():
                if col == sub:
                    colId = getColnum(col, wksheet)
                    wksheet.cell(row=num_row, column=colId).value = scr
        elif col == 'catchement_area':
            colId = getColnum(col, wksheet)
            wksheet.cell(row=num_row, column=colId).value = catchemment_status
        elif col == 'prefered_programme':
            colId = getColnum(col, wksheet)
            wksheet.cell(row=num_row, column=colId).value = prefered_programme
        
        elif col == 'olevel_aggr_score':
            colId = getColnum(col, wksheet)
            wksheet.cell(row=num_row, column=colId).value = olevelAggrScore
        elif col == 'utme_aggr_score':
            colId = getColnum(col, wksheet)
            wksheet.cell(row=num_row, column=colId).value = utmeAggr_Score
        elif col == 'postutme_aggr_score':
            colId = getColnum(col, wksheet)
            wksheet.cell(row=num_row, column=colId).value = posutmeAggrscore
        elif col == 'total_aggregate_score':
            colId = getColnum(col, wksheet)
            wksheet.cell(row=num_row, column=colId).value = final_score
        elif col == 'student_age':
            colId = getColnum(col, wksheet)
            wksheet.cell(row=num_row, column=colId).value = st_age
        elif col == 'admission_status':
            if math_status and eng_status and admission_status:
                    math_eng_status = True
            else:
                math_eng_status = False
            
            colId = getColnum(col, wksheet)
            wksheet.cell(row=num_row, column=colId).value = math_eng_status
        
        elif col == 'sex':
            colId = getColnum(col, wksheet)
            wksheet.cell(row=num_row, column=colId).value = gender_status
        elif col == 'first_choice_status':
            colId = getColnum(col, wksheet)
            wksheet.cell(row=num_row, column=colId).value = first_choice_status
        elif col == 'student_ranking':
            colId = getColnum(col, wksheet)
            if key == studentID:
                wksheet.cell(row=num_row, column=colId).value = student_ranking
        elif col == 'offered_programme':
            recom_programme =   OfferedProgramme(prefered_programme, math_eng_status, student_ranking, programme_max_quota)
            colId = getColnum(col, wksheet)
            wksheet.cell(row=num_row, column=colId).value = recom_programme.return_offered_programme()